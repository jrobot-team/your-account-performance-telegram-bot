import time
import datetime
from operator import itemgetter

import openpyxl
import requests
import xlsxwriter
import pymysql.cursors
import pandas as pd

import config


def average_price_of_inventory(transactions):
	""" Calculates the average price of stock.
	The FIFO method is used as the most common method for valuing stock.This
	method assumes that the first inventories bought are the first ones to
	be sold.

	Arguments:
	transactions: pandas.DataFrame, ordered by index, with a positive VOLUME
	if a stock is bought, and with a negative one if sold.
	Example of transactions:
		+------------+--------+---------+
		|    index   |  PRICE |  VOLUME |
		+------------+--------+---------+
		|'2017-10-16'| 197.67 |    20   |
		|'2018-02-07'| 263.30 |   -10   |
		+------------+--------+---------+

	The function returns a float number.

	"""

	# Closing value of inventory
	closing_value = 0
	# Closing volume of inventory.
	# The sign of closing volume defines whether a long or short position.
	closing_volume = transactions['VOLUME'].sum()
	# Cumulative volume of recent transactions.
	cumulative_volume = 0

	for index in reversed(transactions.index):
		# FOR loop starts with recent transactions. If they total in a long
		# position, then only buying transactions to be counted as they
		# increase closing volume. If they result in a short, then only
		# selling transactions to be counted as they increase resulting
		# volume.
		if closing_volume > 0:
			if transactions.at[index, 'VOLUME'] < 0:
				recent_volume = 0
			else:
				recent_volume = transactions.at[index, 'VOLUME']
		else:
			if transactions.at[index, 'VOLUME'] > 0:
				recent_volume = 0
			else:
				recent_volume = transactions.at[index, 'VOLUME']
		# Summing only the units left in the closing inventory.
		if closing_volume > 0:
			volumes_left = max(0, min(recent_volume,
									  closing_volume - cumulative_volume))
		else:
			volumes_left = min(0, max(recent_volume,
									  closing_volume - cumulative_volume))
		cumulative_volume += recent_volume
		# Summing only the values of units left in the closing inventory.
		closing_value += transactions.at[index, 'PRICE'] * volumes_left
	# Returns average price of units left in closing inventory
	return closing_value / closing_volume


class DataBase:
	"""
	DataBase Class
	"""

	@staticmethod
	def deploy_database():
		connection = pymysql.connect(
			host=config.db_host,
			user=config.db_user,
			password=config.db_password,
			db=config.db_database,
			charset=config.db_charset,
			cursorclass=pymysql.cursors.DictCursor)
		try:
			with connection.cursor() as cursor:
				sql = '''
				CREATE TABLE IF NOT EXISTS `accountamount` (
					`id` int(11) NOT NULL AUTO_INCREMENT,
					`uid` int(11) COLLATE utf8_general_ci NOT NULL,
					`date` varchar(255) COLLATE utf8_general_ci NOT NULL,
					`input_date` varchar(255) COLLATE utf8_general_ci NOT NULL,
					`amount` varchar(255) COLLATE utf8_general_ci NOT NULL,
					`broker` varchar(255) COLLATE utf8_general_ci,
					PRIMARY KEY (`id`)
				) ENGINE=InnoDB DEFAULT CHARSET=utf8 COLLATE=utf8_general_ci
				AUTO_INCREMENT=1;
				'''
				cursor.execute(sql)
				sql = '''
				CREATE TABLE IF NOT EXISTS `accountminusamount` (
					`id` int(11) NOT NULL AUTO_INCREMENT,
					`uid` int(11) COLLATE utf8_general_ci NOT NULL,
					`date` varchar(255) COLLATE utf8_general_ci NOT NULL,
					`input_date` varchar(255) COLLATE utf8_general_ci NOT NULL,
					`amount` varchar(255) COLLATE utf8_general_ci NOT NULL,
					`broker` varchar(255) COLLATE utf8_general_ci,
					PRIMARY KEY (`id`)
				) ENGINE=InnoDB DEFAULT CHARSET=utf8 COLLATE=utf8_general_ci
				AUTO_INCREMENT=1;
				'''
				cursor.execute(sql)
				sql = '''
				CREATE TABLE IF NOT EXISTS `buystock` (
					`id` int(11) NOT NULL AUTO_INCREMENT,
					`uid` int(11) COLLATE utf8_general_ci NOT NULL,
					`date` varchar(255) COLLATE utf8_general_ci NOT NULL,
					`input_date` varchar(255) COLLATE utf8_general_ci NOT NULL,
					`ticker` varchar(255) COLLATE utf8_general_ci NOT NULL,
					`count` int(11) COLLATE utf8_general_ci NOT NULL,
					`broker` varchar(255) COLLATE utf8_general_ci NOT NULL,
					`price` varchar(255) COLLATE utf8_general_ci NOT NULL,
					`api_price` varchar(255) COLLATE utf8_general_ci NOT NULL,
					PRIMARY KEY (`id`)
				) ENGINE=InnoDB DEFAULT CHARSET=utf8 COLLATE=utf8_general_ci
				AUTO_INCREMENT=1;
				'''
				cursor.execute(sql)
				sql = '''
				CREATE TABLE IF NOT EXISTS `salestock` (
					`id` int(11) NOT NULL AUTO_INCREMENT,
					`uid` int(11) COLLATE utf8_general_ci NOT NULL,
					`date` varchar(255) COLLATE utf8_general_ci NOT NULL,
					`input_date` varchar(255) COLLATE utf8_general_ci NOT NULL,
					`ticker` varchar(255) COLLATE utf8_general_ci NOT NULL,
					`count` int(11) COLLATE utf8_general_ci NOT NULL,
					`broker` varchar(255) COLLATE utf8_general_ci NOT NULL,
					`price` varchar(255) COLLATE utf8_general_ci NOT NULL,
					PRIMARY KEY (`id`)
				) ENGINE=InnoDB DEFAULT CHARSET=utf8 COLLATE=utf8_general_ci
				AUTO_INCREMENT=1;
				'''
				cursor.execute(sql)
				sql = '''
				CREATE TABLE IF NOT EXISTS `buybond` (
					`id` int(11) NOT NULL AUTO_INCREMENT,
					`uid` int(11) COLLATE utf8_general_ci NOT NULL,
					`date` varchar(255) COLLATE utf8_general_ci NOT NULL,
					`input_date` varchar(255) COLLATE utf8_general_ci NOT NULL,
					`ticker` varchar(255) COLLATE utf8_general_ci NOT NULL,
					`count` int(11) COLLATE utf8_general_ci NOT NULL,
					`nkd` varchar(255) COLLATE utf8_general_ci NOT NULL,
					`broker` varchar(255) COLLATE utf8_general_ci,
					`price` varchar(255) COLLATE utf8_general_ci NOT NULL,
					`api_price` varchar(255) COLLATE utf8_general_ci NOT NULL,
					`api_nkd` varchar(255) COLLATE utf8_general_ci NOT NULL,
					`api_FACEVALUE` varchar(255) COLLATE utf8_general_ci NOT NULL,
					`api_ACCINT` varchar(255) COLLATE utf8_general_ci NOT NULL,
					PRIMARY KEY (`id`)
				) ENGINE=InnoDB DEFAULT CHARSET=utf8 COLLATE=utf8_general_ci
				AUTO_INCREMENT=1;
				'''
				cursor.execute(sql)
				sql = '''
				CREATE TABLE IF NOT EXISTS `salebond` (
					`id` int(11) NOT NULL AUTO_INCREMENT,
					`uid` int(11) COLLATE utf8_general_ci NOT NULL,
					`date` varchar(255) COLLATE utf8_general_ci NOT NULL,
					`input_date` varchar(255) COLLATE utf8_general_ci NOT NULL,
					`ticker` varchar(255) COLLATE utf8_general_ci NOT NULL,
					`count` int(11) COLLATE utf8_general_ci NOT NULL,
					`broker` varchar(255) COLLATE utf8_general_ci,
					`nkd` varchar(255) COLLATE utf8_general_ci NOT NULL,
					`price` varchar(255) COLLATE utf8_general_ci NOT NULL,
					PRIMARY KEY (`id`)
				) ENGINE=InnoDB DEFAULT CHARSET=utf8 COLLATE=utf8_general_ci
				AUTO_INCREMENT=1;
				'''
				cursor.execute(sql)
				sql = '''
				CREATE TABLE IF NOT EXISTS `taxes` (
					`id` int(11) NOT NULL AUTO_INCREMENT,
					`uid` int(11) COLLATE utf8_general_ci NOT NULL,
					`date` varchar(255) COLLATE utf8_general_ci NOT NULL,
					`input_date` varchar(255) COLLATE utf8_general_ci NOT NULL,
					`amount` varchar(255) COLLATE utf8_general_ci NOT NULL,
					`broker` varchar(255) COLLATE utf8_general_ci,
					PRIMARY KEY (`id`)
				) ENGINE=InnoDB DEFAULT CHARSET=utf8 COLLATE=utf8_general_ci
				AUTO_INCREMENT=1;
				'''
				cursor.execute(sql)
				sql = '''
				CREATE TABLE IF NOT EXISTS `comissions` (
					`id` int(11) NOT NULL AUTO_INCREMENT,
					`uid` int(11) COLLATE utf8_general_ci NOT NULL,
					`date` varchar(255) COLLATE utf8_general_ci NOT NULL,
					`input_date` varchar(255) COLLATE utf8_general_ci NOT NULL,
					`amount` varchar(255) COLLATE utf8_general_ci NOT NULL,
					`broker` varchar(255) COLLATE utf8_general_ci,
					PRIMARY KEY (`id`)
				) ENGINE=InnoDB DEFAULT CHARSET=utf8 COLLATE=utf8_general_ci
				AUTO_INCREMENT=1;
				'''
				cursor.execute(sql)
				sql = '''
				CREATE TABLE IF NOT EXISTS `couponincome` (
					`id` int(11) NOT NULL AUTO_INCREMENT,
					`uid` int(11) COLLATE utf8_general_ci NOT NULL,
					`date` varchar(255) COLLATE utf8_general_ci NOT NULL,
					`input_date` varchar(255) COLLATE utf8_general_ci NOT NULL,
					`bond` varchar(255) COLLATE utf8_general_ci NOT NULL,
					`amount` varchar(255) COLLATE utf8_general_ci NOT NULL,
					`broker` varchar(255) COLLATE utf8_general_ci,
					PRIMARY KEY (`id`)
				) ENGINE=InnoDB DEFAULT CHARSET=utf8 COLLATE=utf8_general_ci
				AUTO_INCREMENT=1;
				'''
				cursor.execute(sql)
				sql = '''
				CREATE TABLE IF NOT EXISTS `dividends` (
					`id` int(11) NOT NULL AUTO_INCREMENT,
					`uid` int(11) COLLATE utf8_general_ci NOT NULL,
					`date` varchar(255) COLLATE utf8_general_ci NOT NULL,
					`input_date` varchar(255) COLLATE utf8_general_ci NOT NULL,
					`dividend` varchar(255) COLLATE utf8_general_ci NOT NULL,
					`amount` varchar(255) COLLATE utf8_general_ci NOT NULL,
					`broker` varchar(255) COLLATE utf8_general_ci,
					PRIMARY KEY (`id`)
				) ENGINE=InnoDB DEFAULT CHARSET=utf8 COLLATE=utf8_general_ci
				AUTO_INCREMENT=1;
				'''
				cursor.execute(sql)
			connection.commit()
		finally:
			connection.close()

	@staticmethod
	def add_new_amount(uid, date, input_date, amount, broker):
		"""
		Пополнить счет
		"""
		connection = pymysql.connect(
			host=config.db_host,
			user=config.db_user,
			password=config.db_password,
			db=config.db_database,
			charset=config.db_charset,
			cursorclass=pymysql.cursors.DictCursor)
		try:
			with connection.cursor() as cursor:
				sql = 'INSERT INTO `accountamount` (`uid`, `date`, `input_date`, `amount`, `broker`) VALUES (%s, %s, %s, %s, %s)'
				cursor.execute(sql, (uid, date, input_date, amount, broker))
			connection.commit()
		finally:
			connection.close()

	@staticmethod
	def add_minus_amount(uid, date, input_date, amount, broker):
		"""
		Вывод средств
		"""
		connection = pymysql.connect(
			host=config.db_host,
			user=config.db_user,
			password=config.db_password,
			db=config.db_database,
			charset=config.db_charset,
			cursorclass=pymysql.cursors.DictCursor)
		try:
			with connection.cursor() as cursor:
				sql = 'INSERT INTO `accountminusamount` (`uid`, `date`, `input_date`, `amount`, `broker`) VALUES (%s, %s, %s, %s, %s)'
				cursor.execute(sql, (uid, date, input_date, amount, broker))
			connection.commit()
		finally:
			connection.close()
	
	@staticmethod
	def add_buystock(uid, date, input_date, ticker, count, broker, price, api_price=0):
		"""
		Покупка акций
		"""
		connection = pymysql.connect(
			host=config.db_host,
			user=config.db_user,
			password=config.db_password,
			db=config.db_database,
			charset=config.db_charset,
			cursorclass=pymysql.cursors.DictCursor)
		try:
			with connection.cursor() as cursor:
				sql = 'INSERT INTO `buystock` (`uid`, `date`, `input_date`, `ticker`, `count`, `broker`, `price`, `api_price`) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)'
				cursor.execute(sql, (uid, date, input_date, ticker, count, broker, price, api_price))
			connection.commit()
		finally:
			connection.close()

	@staticmethod
	def add_salestock(uid, date, input_date, ticker, count, broker, price):
		"""
		Продажа акций
		"""
		connection = pymysql.connect(
			host=config.db_host,
			user=config.db_user,
			password=config.db_password,
			db=config.db_database,
			charset=config.db_charset,
			cursorclass=pymysql.cursors.DictCursor)
		try:
			with connection.cursor() as cursor:
				sql = 'INSERT INTO `salestock` (`uid`, `date`, `input_date`, `ticker`, `count`, `broker`, `price`) VALUES (%s, %s, %s, %s, %s, %s, %s)'
				cursor.execute(sql, (uid, date, input_date, ticker, count, broker, price))
			connection.commit()
		finally:
			connection.close()

	@staticmethod
	def add_buybond(uid, date, input_date, ticker, count, nkd, price, broker, api_price, api_nkd, api_FACEVALUE, api_ACCINT):
		"""
		Покупка облигаций
		"""
		connection = pymysql.connect(
			host=config.db_host,
			user=config.db_user,
			password=config.db_password,
			db=config.db_database,
			charset=config.db_charset,
			cursorclass=pymysql.cursors.DictCursor)
		try:
			with connection.cursor() as cursor:
				sql = 'INSERT INTO `buybond` (`uid`, `date`, `input_date` ,`ticker`, `count`, `nkd`, `price`, `broker`, `api_price`, `api_nkd`, `api_FACEVALUE`, `api_ACCINT`) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)'
				cursor.execute(sql, (uid, date, input_date, ticker, count, nkd, price, broker, api_price, api_nkd, api_FACEVALUE, api_ACCINT))
			connection.commit()
		finally:
			connection.close()

	@staticmethod
	def add_salebond(uid, date, input_date, ticker, count, broker, nkd, price):
		"""
		Продажа облигации
		"""
		connection = pymysql.connect(
			host=config.db_host,
			user=config.db_user,
			password=config.db_password,
			db=config.db_database,
			charset=config.db_charset,
			cursorclass=pymysql.cursors.DictCursor)
		try:
			with connection.cursor() as cursor:
				sql = 'INSERT INTO `salebond` (`uid`, `date`, `input_date`, `ticker`, `count`, `broker`, `nkd`, `price`) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)'
				cursor.execute(sql, (uid, date, input_date, ticker, count, broker, nkd, price))
			connection.commit()
		finally:
			connection.close()

	@staticmethod
	def add_new_tax(uid, date, input_date, amount, broker):
		"""
		Налог
		"""
		connection = pymysql.connect(
			host=config.db_host,
			user=config.db_user,
			password=config.db_password,
			db=config.db_database,
			charset=config.db_charset,
			cursorclass=pymysql.cursors.DictCursor)
		try:
			with connection.cursor() as cursor:
				sql = 'INSERT INTO `taxes` (`uid`, `date`, `input_date`, `amount`, `broker`) VALUES (%s, %s, %s, %s, %s)'
				cursor.execute(sql, (uid, date, input_date, amount, broker))
			connection.commit()
		finally:
			connection.close()
	
	@staticmethod
	def add_new_commission(uid, date, input_date, amount, broker):
		"""
		Комиссия
		"""
		connection = pymysql.connect(
			host=config.db_host,
			user=config.db_user,
			password=config.db_password,
			db=config.db_database,
			charset=config.db_charset,
			cursorclass=pymysql.cursors.DictCursor)
		try:
			with connection.cursor() as cursor:
				sql = 'INSERT INTO `comissions` (`uid`, `date`, `input_date`, `amount`, `broker`) VALUES (%s, %s, %s, %s, %s)'
				cursor.execute(sql, (uid, date, input_date, amount, broker))
			connection.commit()
		finally:
			connection.close()

	@staticmethod
	def add_new_couponincome(uid, date, input_date, bond, amount, broker):
		"""
		Купонный доход
		"""
		connection = pymysql.connect(
			host=config.db_host,
			user=config.db_user,
			password=config.db_password,
			db=config.db_database,
			charset=config.db_charset,
			cursorclass=pymysql.cursors.DictCursor)
		try:
			with connection.cursor() as cursor:
				sql = 'INSERT INTO `couponincome` (`uid`, `date`, `input_date`, `bond`, `amount`, `broker`) VALUES (%s, %s, %s, %s, %s, %s)'
				cursor.execute(sql, (uid, date, input_date, bond, amount, broker))
			connection.commit()
		finally:
			connection.close()
	
	@staticmethod
	def add_new_dividend(uid, date, input_date, dividend, amount, broker):
		"""
		Зачисление дивидендов
		"""
		connection = pymysql.connect(
			host=config.db_host,
			user=config.db_user,
			password=config.db_password,
			db=config.db_database,
			charset=config.db_charset,
			cursorclass=pymysql.cursors.DictCursor)
		try:
			with connection.cursor() as cursor:
				sql = 'INSERT INTO `dividends` (`uid`, `date`, `input_date`, `dividend`, `amount`, `broker`) VALUES (%s, %s, %s, %s, %s, %s)'
				cursor.execute(sql, (uid, date, input_date, dividend, amount, broker))
			connection.commit()
		finally:
			connection.close()

	@staticmethod
	def delete_operation(table, uid, id):
		"""
		Удалить операцию
		"""
		connection = pymysql.connect(
			host=config.db_host,
			user=config.db_user,
			password=config.db_password,
			db=config.db_database,
			charset=config.db_charset,
			cursorclass=pymysql.cursors.DictCursor)
		try:
			with connection.cursor() as cursor:
				sql = 'DELETE FROM {!s} WHERE uid=%s AND id=%s' .format(table)
				cursor.execute(sql, (uid, id))
			connection.commit()
		finally:
			connection.close()
	
	@staticmethod
	def get_operation(table, uid, id):
		"""
		Получить операцию
		"""
		connection = pymysql.connect(
			host=config.db_host,
			user=config.db_user,
			password=config.db_password,
			db=config.db_database,
			charset=config.db_charset,
			cursorclass=pymysql.cursors.DictCursor)
		try:
			with connection.cursor() as cursor:
				sql = 'SELECT * FROM {!s} WHERE uid=%s AND id=%s' .format(table)
				cursor.execute(sql, (uid, id))
				res = cursor.fetchone()
				return res
			connection.commit()
		except:
			return None
		finally:
			connection.close()
		return None


def get_history(uid, start_timestamp, end_timestamp):
	"""
	Получить историю операций в виде сообщений
	"""
	connection = pymysql.connect(
		host=config.db_host,
		user=config.db_user,
		password=config.db_password,
		db=config.db_database,
		charset=config.db_charset,
		cursorclass=pymysql.cursors.DictCursor)
	try:
		operations = []
		with connection.cursor() as cursor:

			sql = 'SELECT * FROM accountamount WHERE uid=%s AND input_date < %s AND input_date > %s'
			cursor.execute(sql, (uid, start_timestamp, end_timestamp))
			res = cursor.fetchall()
			for x in res:
				operations.append(x)
				operations[-1]['title'] = 'Пополнение счета'
				operations[-1]['table'] = 'accountamount'
			
			sql = 'SELECT * FROM accountminusamount WHERE uid=%s AND input_date < %s AND input_date > %s'
			cursor.execute(sql, (uid, start_timestamp, end_timestamp))
			res = cursor.fetchall()
			for x in res:
				operations.append(x)
				operations[-1]['title'] = 'Вывод средств'
				operations[-1]['table'] = 'accountminusamount'
			
			sql = 'SELECT * FROM buystock WHERE uid=%s AND input_date < %s AND input_date > %s'
			cursor.execute(sql, (uid, start_timestamp, end_timestamp))
			res = cursor.fetchall()
			for x in res:
				operations.append(x)
				operations[-1]['title'] = 'Покупка акций'
				operations[-1]['table'] = 'buystock'
			
			sql = 'SELECT * FROM salestock WHERE uid=%s AND input_date < %s AND input_date > %s'
			cursor.execute(sql, (uid, start_timestamp, end_timestamp))
			res = cursor.fetchall()
			for x in res:
				operations.append(x)
				operations[-1]['title'] = 'Продажа акций'
				operations[-1]['table'] = 'salestock'
			
			sql = 'SELECT * FROM buybond WHERE uid=%s AND input_date < %s AND input_date > %s'
			cursor.execute(sql, (uid, start_timestamp, end_timestamp))
			res = cursor.fetchall()
			for x in res:
				operations.append(x)
				operations[-1]['title'] = 'Покупка облигаций'
				operations[-1]['table'] = 'buybond'
			
			sql = 'SELECT * FROM salebond WHERE uid=%s AND input_date < %s AND input_date > %s'
			cursor.execute(sql, (uid, start_timestamp, end_timestamp))
			res = cursor.fetchall()
			for x in res:
				operations.append(x)
				operations[-1]['title'] = 'Продажа облигаций'
				operations[-1]['table'] = 'salebond'
			
			sql = 'SELECT * FROM taxes WHERE uid=%s AND input_date < %s AND input_date > %s'
			cursor.execute(sql, (uid, start_timestamp, end_timestamp))
			res = cursor.fetchall()
			for x in res:
				operations.append(x)
				operations[-1]['title'] = 'Удержание налога'
				operations[-1]['table'] = 'taxes'
			
			sql = 'SELECT * FROM comissions WHERE uid=%s AND input_date < %s AND input_date > %s'
			cursor.execute(sql, (uid, start_timestamp, end_timestamp))
			res = cursor.fetchall()
			for x in res:
				operations.append(x)
				operations[-1]['title'] = 'Удержание комиссии'
				operations[-1]['table'] = 'comissions'
			
			sql = 'SELECT * FROM couponincome WHERE uid=%s AND input_date < %s AND input_date > %s'
			cursor.execute(sql, (uid, start_timestamp, end_timestamp))
			res = cursor.fetchall()
			for x in res:
				operations.append(x)
				operations[-1]['title'] = 'Зачисление купонного дохода'
				operations[-1]['table'] = 'couponincome'
			
			sql = 'SELECT * FROM dividends WHERE uid=%s AND input_date < %s AND input_date > %s'
			cursor.execute(sql, (uid, start_timestamp, end_timestamp))
			res = cursor.fetchall()
			for x in res:
				operations.append(x)
				operations[-1]['title'] = 'Зачисление дивидендов'
				operations[-1]['table'] = 'dividends'

		connection.commit()
		return sorted(operations, key=itemgetter('input_date'), reverse=False)
	finally:
		connection.close()


class Moex:
	"""
	API биржи Moex
	"""

	@staticmethod
	def get_stock_price(tiker):
		"""
		Получить стоимость акции
		"""
		try:
			current_date = datetime.datetime.now().strftime('%Y-%m-%d')
			last_day = (datetime.datetime.now() - datetime.timedelta(days=30)).strftime('%Y-%m-%d')
			res = requests.get('http://iss.moex.com/iss/securities.json?q={!s}'.format(tiker))
			board = ''
			for x in res.json()['securities']['data']:
				if x[1] == tiker:
					board = x[-1]
					break
			if len(board) == 0:
				return None
			# board = res.json()['securities']['data'][0][-1]
			url = 'http://iss.moex.com/iss/history/engines/stock/markets/shares/boards/{!s}/securities/{!s}.json?from={!s}&till={!s}'.format(
				board, tiker, last_day, current_date
			)
			print(url)
			res = requests.get(url)
			price = res.json()['history']['data'][-1][9]
			return float('{0: >#016.4f}'.format(float(price)).strip())
		except Exception as e:
			print(e)
			return None

	@staticmethod
	def get_bond_price(code):
		"""
		DEPRECATED!
		Получить стоимость облигации
		"""
		try:
			current_date = datetime.datetime.now().strftime('%Y-%m-%d')
			last_day = (datetime.datetime.now() - datetime.timedelta(days=30)).strftime('%Y-%m-%d')
			res = requests.get('http://iss.moex.com/iss/securities.json?q={!s}'.format(code))
			board = res.json()['securities']['data'][0][-1]
			url = 'http://iss.moex.com/iss/history/engines/stock/markets/bonds/boards/{!s}/securities/{!s}.json?from={!s}&till={!s}'.format(
				board, code, last_day, current_date
			)
			res = requests.get(url)
			print(url)
			price = res.json()['history']['data'][-1][9]
			return float('{0: >#016.4f}'.format(float(price)).strip())
		except Exception as e:
			print(e)
			return None

	@staticmethod
	def get_bond_nkd(code):
		"""
		DEPRECATED!
		Получить стоимость облигации
		"""
		try:
			current_date = datetime.datetime.now().strftime('%Y-%m-%d')
			last_day = (datetime.datetime.now() - datetime.timedelta(days=30)).strftime('%Y-%m-%d')
			res = requests.get('http://iss.moex.com/iss/securities.json?q={!s}'.format(code))
			board = res.json()['securities']['data'][0][-1]
			url = 'http://iss.moex.com/iss/history/engines/stock/markets/bonds/boards/{!s}/securities/{!s}.json?from={!s}&till={!s}'.format(
				board, code, last_day, current_date
			)
			res = requests.get(url)
			print(url)
			nkd = res.json()['history']['data'][-1][27]
			return float('{0: >#016.4f}'.format(float(nkd)).strip())
		except Exception as e:
			print(e)
			return None

	@staticmethod
	def get_bond_data(code):
		"""
		Получить стоимость облигации
		"""
		days = 60
		while True:
			if days > 1740:
				return None
			try:
				current_date = datetime.datetime.now().strftime('%Y-%m-%d')
				last_day = (datetime.datetime.now() - datetime.timedelta(days=days)).strftime('%Y-%m-%d')
				res = requests.get('http://iss.moex.com/iss/securities.json?q={!s}'.format(code))
				board = res.json()['securities']['data'][0][-1]
				url = 'http://iss.moex.com/iss/history/engines/stock/markets/bonds/boards/{!s}/securities/{!s}.json?from={!s}&till={!s}'.format(
					board, code, last_day, current_date
				)
				res = requests.get(url)
				print(url)

				price = res.json()['history']['data'][-1][9]
				nkd = res.json()['history']['data'][-1][27]
				FACEVALUE = res.json()['history']['data'][-1][30]
				ACCINT = res.json()['history']['data'][-1][10]

				if nkd is None:
					nkd = 0

				price = float('{0: >#016.4f}'.format(float(price)).strip())
				nkd = float('{0: >#016.4f}'.format(float(nkd)).strip())
				FACEVALUE = float('{0: >#016.4f}'.format(float(FACEVALUE)).strip())
				ACCINT = float('{0: >#016.4f}'.format(float(ACCINT)).strip())

				return {
					'price': price,
					'nkd': nkd,
					'FACEVALUE': FACEVALUE,
					'ACCINT': ACCINT,
				}
			except IndexError as e:
				print(e)
				days += 60
				continue
			except Exception as e:
				print(e)
				return None


def update_moex():
	"""
	Обновить стоимости акций и облигаций
	"""
	connection = pymysql.connect(
		host=config.db_host,
		user=config.db_user,
		password=config.db_password,
		db=config.db_database,
		charset=config.db_charset,
		cursorclass=pymysql.cursors.DictCursor)
	try:
		with connection.cursor() as cursor:
			# Обработать акции
			sql = 'SELECT * FROM buystock'
			cursor.execute(sql)
			res = cursor.fetchall()
			print(res)
			for stock in res:
				new_api_price = Moex.get_stock_price(stock['ticker'])
				sql = 'UPDATE buystock SET api_price=%s WHERE id=%s'
				cursor.execute(sql, (new_api_price, stock['id']))
			# Обработать облигации
			sql = 'SELECT * FROM buybond'
			cursor.execute(sql)
			res = cursor.fetchall()
			print(res)
			for bond in res:
				data = Moex.get_bond_data(bond['ticker'])
				sql = 'UPDATE buybond SET api_price=%s, api_nkd=%s, api_FACEVALUE=%s, api_ACCINT=%s WHERE id=%s'
				cursor.execute(sql, (data['price'], data['nkd'], data['FACEVALUE'], data['ACCINT'], bond['id']))
		connection.commit()
	finally:
		connection.close()
	return None


def get_timestamp(str_date):
	"""
	Получить timestamp из строка 

	:param: str_date - строка формата ДД.ММ.ГГГГ
	:return: timestamp (int) если строка валидная или None
	"""
	date_arr = str_date.split('.')
	if not len(date_arr) == 3:
		return None
	try:
		dt = datetime.datetime(int(date_arr[2]), int(date_arr[1]), int(date_arr[0])) 
		# Добавить +3 часа, что бы сойтись с мировым временем
		dt += datetime.timedelta(hours=3)
	except:
		return None
	try:
		unixtime = int(time.mktime(dt.timetuple()))
	except:
		return None
	return unixtime


def get_portfolio(uid):
	"""
	Получить информацию о портфеле пользователя
	"""
	connection = pymysql.connect(
		host=config.db_host,
		user=config.db_user,
		password=config.db_password,
		db=config.db_database,
		charset=config.db_charset,
		cursorclass=pymysql.cursors.DictCursor)
	try:
		result = {
			'stocks': [],
			'bonds': [],
		}
		with connection.cursor() as cursor:
			# Получить все купленные тикеры акций
			sql = 'SELECT DISTINCT ticker FROM buystock WHERE uid=%s'
			cursor.execute(sql, (uid,))
			tickers_buy = cursor.fetchall()
			sql = 'SELECT DISTINCT ticker FROM salestock WHERE uid=%s'
			cursor.execute(sql, (uid,))
			tickers_sale = cursor.fetchall()
			tickers = []
			for x in tickers_buy:
				if x['ticker'] not in tickers:
					tickers.append(x['ticker'])
			for x in tickers_sale:
				if x['ticker'] not in tickers:
					tickers.append(x['ticker'])
			print(tickers)
			for ticker in tickers:
				stock_arr = []
				date_arr = []
				sql = 'SELECT * FROM buystock WHERE uid=%s AND ticker=%s'
				cursor.execute(sql, (uid, ticker))
				res1 = cursor.fetchall()
				amount = 0
				buy_count = 0
				for x in res1:
					amount += float(x['price']) * int(x['count'])
					buy_count += x['count']
					stock_arr.append([float(x['price']), int(x['count'])])
					# date_arr.append(datetime.datetime.utcfromtimestamp(int(x['date'])).strftime('%Y-%m-%d'))
					date_arr.append(int(x['date']))
				sql = 'SELECT * FROM salestock WHERE uid=%s AND ticker=%s'
				cursor.execute(sql, (uid, ticker))
				res2 = cursor.fetchall()
				sale_count = 0
				for x in res2:
					amount -= float(x['price']) * int(x['count'])
					sale_count += x['count']
					stock_arr.append([float(x['price']), int(x['count']) * -1])
					# date_arr.append(datetime.datetime.utcfromtimestamp(int(x['date'])).strftime('%Y-%m-%d'))
					date_arr.append(int(x['date']))
				# Количество акций по тикеру
				count = buy_count - sale_count
				if count == 0:
					continue

				date_arr = tuple(date_arr)
				transactions = pd.DataFrame(stock_arr, columns=['PRICE', 'VOLUME'], index=date_arr)
				average_price = average_price_of_inventory(transactions)

				# Текущая стоимость
				if len(res1) == 0:
					api_price = Moex.get_stock_price(ticker)
				else:
					api_price = res1[0]['api_price']
				current_price = float(api_price) * count
				# Разница стоимости
				price_difference = current_price - average_price * count
				result['stocks'].append({
					'ticker': ticker,
					'count': count,
					'average_price': float('{0: >#016.4f}'.format(float(average_price))),
					'close_price': float('{0: >#016.4f}'.format(float(api_price))),
					'current_price': float('{0: >#016.4f}'.format(float(current_price))),
					'price_difference': float('{0: >#016.4f}'.format(float(price_difference))),
				})

			# Получить все купленные тикеры облигаций
			sql = 'SELECT DISTINCT ticker FROM buybond WHERE uid=%s'
			cursor.execute(sql, (uid,))
			tickers_buy = cursor.fetchall()
			sql = 'SELECT DISTINCT ticker FROM salebond WHERE uid=%s'
			cursor.execute(sql, (uid,))
			tickers_sale = cursor.fetchall()
			tickers = []
			for x in tickers_buy:
				if x['ticker'] not in tickers:
					tickers.append(x['ticker'])
			for x in tickers_sale:
				if x['ticker'] not in tickers:
					tickers.append(x['ticker'])
			print(tickers)
			for ticker in tickers:
				sql = 'SELECT * FROM buybond WHERE uid=%s AND ticker=%s'
				cursor.execute(sql, (uid, ticker))
				res1 = cursor.fetchall()
				amount = 0
				buy_count = 0
				for x in res1:
					amount += (float(x['price']) + float(x['nkd'])) * int(x['count'])
					buy_count += x['count']
				sql = 'SELECT * FROM salebond WHERE uid=%s AND ticker=%s'
				cursor.execute(sql, (uid, ticker))
				res2 = cursor.fetchall()
				sale_count = 0
				for x in res2:
					amount -= (float(x['price']) + float(x['nkd'])) * int(x['count'])
					sale_count += x['count']
				# Количество облигаций по тикеру
				count = buy_count - sale_count
				if count == 0:
					continue

				# Сумма всех цен операций покупки
				# Средняя стоимость
				average_price = amount / count

				# Текущая стоимость
				if len(res1) == 0:
					data = Moex.get_bond_data(ticker)
					api_price = data['price']
					api_nkd = data['nkd']
					api_FACEVALUE = data['FACEVALUE']
					api_ACCINT = data['ACCINT']
				else:
					api_price = res1[0]['api_price']
					api_nkd = res1[0]['api_nkd']
					api_FACEVALUE = res1[0]['api_FACEVALUE']
					api_ACCINT = res1[0]['api_ACCINT']
				# current_price = (float(api_price) + float(api_nkd)) * count
				current_price = (float(api_price) * float(api_FACEVALUE) / 100 + float(api_ACCINT)) * count
				# Разница стоимости
				price_difference = current_price - average_price * count
				# price_difference = (average_price - current_price) * count
				result['bonds'].append({
					'ticker': ticker,
					'count': count,
					'average_price': float('{0: >#016.4f}'.format(float(average_price))),
					'close_price': float('{0: >#016.4f}'.format(float(api_price))),
					'current_price': float('{0: >#016.4f}'.format(float(current_price))),
					'price_difference': float('{0: >#016.4f}'.format(float(price_difference))),
				})
		connection.commit()
		return result
	finally:
		connection.close()


def get_portfolio_amount(uid):
	"""
	Получить накопленную прибыль/убыток
	"""
	data = get_portfolio(uid)
	print(data)
	amount = 0
	for x in data['stocks']:
		amount += x['price_difference']
	for x in data['bonds']:
		amount += x['price_difference']
	return float(amount)


def get_account_state(uid):
	"""
	Получить состояние счета пользователя
	"""
	connection = pymysql.connect(
		host=config.db_host,
		user=config.db_user,
		password=config.db_password,
		db=config.db_database,
		charset=config.db_charset,
		cursorclass=pymysql.cursors.DictCursor)
	try:
		broker_amount = 0
		money_amount = 0
		amount = 0
		with connection.cursor() as cursor:
			# Высчитать пополнение счета
			sql = 'SELECT * FROM accountamount WHERE uid=%s'
			cursor.execute(sql, (uid,))
			res = cursor.fetchall()
			for x in res:
				amount += float(x['amount'])
				money_amount += float(x['amount'])
			# Высчитать вывод средств
			sql = 'SELECT * FROM accountminusamount WHERE uid=%s'
			cursor.execute(sql, (uid,))
			res = cursor.fetchall()
			for x in res:
				amount -= float(x['amount'])
				money_amount -= float(x['amount'])
			# Высчитать покупку акций
			sql = 'SELECT * FROM buystock WHERE uid=%s'
			cursor.execute(sql, (uid,))
			res = cursor.fetchall()
			for x in res:
				amount -= float(x['price']) * int(x['count'])
				broker_amount += float(x['price']) * int(x['count'])
			# Высчитать продажу акций
			sql = 'SELECT * FROM salestock WHERE uid=%s'
			cursor.execute(sql, (uid,))
			res = cursor.fetchall()
			for x in res:
				amount += float(x['price']) * int(x['count'])
				broker_amount -= float(x['price']) * int(x['count'])
			# Высчитать покупку облигаций
			sql = 'SELECT * FROM buybond WHERE uid=%s'
			cursor.execute(sql, (uid,))
			res = cursor.fetchall()
			for x in res:
				amount -= (float(x['price']) + float(x['nkd'])) * int(x['count'])
				broker_amount += (float(x['price']) + float(x['nkd'])) * int(x['count'])
			# Высчитать продажу облигаций
			sql = 'SELECT * FROM salebond WHERE uid=%s'
			cursor.execute(sql, (uid,))
			res = cursor.fetchall()
			for x in res:
				amount += (float(x['price']) + float(x['nkd'])) * int(x['count'])
				broker_amount -= (float(x['price']) + float(x['nkd'])) * int(x['count'])
			# Высчитать удержание налога
			sql = 'SELECT * FROM taxes WHERE uid=%s'
			cursor.execute(sql, (uid,))
			res = cursor.fetchall()
			for x in res:
				amount -= float(x['amount'])
				money_amount -= float(x['amount'])
			# Высчитать удержание комиссии
			sql = 'SELECT * FROM comissions WHERE uid=%s'
			cursor.execute(sql, (uid,))
			res = cursor.fetchall()
			for x in res:
				amount -= float(x['amount'])
				money_amount -= float(x['amount'])
			# Высчитать зачисление купонного дохода
			sql = 'SELECT * FROM couponincome WHERE uid=%s'
			cursor.execute(sql, (uid,))
			res = cursor.fetchall()
			for x in res:
				amount += float(x['amount'])
				money_amount += float(x['amount'])
			# Высчитать стоимость дивидедов
			sql = 'SELECT * FROM dividends WHERE uid=%s'
			cursor.execute(sql, (uid,))
			res = cursor.fetchall()
			for x in res:
				amount += float(x['amount'])
				money_amount += float(x['amount'])
		connection.commit()
		data1 = get_portfolio(uid)
		print(data1)
		amount1 = 0
		for x in data1['stocks']:
			amount1 += x['current_price']
		for x in data1['bonds']:
			amount1 += x['current_price']
		broker_amount = float(amount1) + float(amount)
		amount = float(amount)
		return {
			'amount': amount,
			'money_amount': money_amount,
			'broker_amount': broker_amount,
		}
	finally:
		connection.close()


def standart_int(number):
	"""
	Привести чило к стандартному для пользователя виду
	"""
	def isInt(n):
		return int(n) == float(n)
	if isInt(number):
		n = '{0:,}'.format(number).replace(',', ' ')
		if '.'in n:
			return n[:-2]
		return n
	else:
		n = float('{0: >#16.2f}'.format(number))
		n = '{0:,}'.format(n).replace(',', ' ')
		'''
		if len(n.split('.')[1]) == 1:
			return '{!s}{!s}'.format(n, '000')
		if len(n.split('.')[1]) == 2:
			return '{!s}{!s}'.format(n, '00')
		if len(n.split('.')[1]) == 3:
			return '{!s}{!s}'.format(n, '0')
		'''
		if len(n.split('.')[1]) == 1:
			return '{!s}{!s}'.format(n, '0')
		return n


def create_excel_export_file(uid):
	"""
	Создать файл excel для экспорта операций
	"""
	filename = 'export_{!s}.xlsx'.format(uid)
	workbook = xlsxwriter.Workbook(filename=filename)
	worksheet = workbook.add_worksheet()
	worksheet.set_column('A:A', 15)
	worksheet.set_column('A:B', 20)
	worksheet.set_column('A:C', 15)
	worksheet.set_column('A:D', 15)
	worksheet.set_column('A:E', 15)
	worksheet.set_column('A:F', 15)
	worksheet.set_column('A:G', 15)
	worksheet.set_column('A:H', 15)

	history = get_history(uid, 2171667370, 0)

	row = 0
	bold = workbook.add_format({'bold': True})
	worksheet.write(row, 0, 'Дата операции')
	worksheet.write(row, 1, 'Тип операции')
	worksheet.write(row, 2, 'Брокер')
	worksheet.write(row, 3, 'Тикер')
	worksheet.write(row, 4, 'Количество')
	worksheet.write(row, 5, 'Цена')
	worksheet.write(row, 6, 'Сумма')
	worksheet.write(row, 7, 'НКД')

	for x in history:
		if x['table'] == 'accountamount':
			row += 1
			_date = datetime.datetime.utcfromtimestamp(int(x['input_date'])).strftime('%d.%m.%Y')
			worksheet.write(row, 0, _date)
			worksheet.write(row, 1, 'Пополнение счета')
			worksheet.write(row, 6, x['amount'])
			worksheet.write(row, 2, x['broker'])
		elif x['table'] == 'accountminusamount':
			row += 1
			_date = datetime.datetime.utcfromtimestamp(int(x['input_date'])).strftime('%d.%m.%Y')
			worksheet.write(row, 0, _date)
			worksheet.write(row, 1, 'Вывод средств')
			worksheet.write(row, 6, x['amount'])
			worksheet.write(row, 2, x['broker'])
		elif x['table'] == 'buystock':
			row += 1
			_date = datetime.datetime.utcfromtimestamp(int(x['input_date'])).strftime('%d.%m.%Y')
			worksheet.write(row, 0, _date)
			worksheet.write(row, 1, 'Покупка акций')
			worksheet.write(row, 3, x['ticker'])
			worksheet.write(row, 4, x['count'])
			worksheet.write(row, 2, x['broker'])
			worksheet.write(row, 5, x['price'])
		elif x['table'] == 'salestock':
			row += 1
			_date = datetime.datetime.utcfromtimestamp(int(x['input_date'])).strftime('%d.%m.%Y')
			worksheet.write(row, 0, _date)
			worksheet.write(row, 1, 'Продажа акций')
			worksheet.write(row, 3, x['ticker'])
			worksheet.write(row, 4, x['count'])
			worksheet.write(row, 2, x['broker'])
			worksheet.write(row, 5, x['price'])
		elif x['table'] == 'buybond':
			row += 1
			_date = datetime.datetime.utcfromtimestamp(int(x['input_date'])).strftime('%d.%m.%Y')
			worksheet.write(row, 0, _date)
			worksheet.write(row, 1, 'Покупка облигаций')
			worksheet.write(row, 3, x['ticker'])
			worksheet.write(row, 4, x['count'])
			worksheet.write(row, 2, x['broker'])
			worksheet.write(row, 5, x['price'])
			worksheet.write(row, 7, x['nkd'])
		elif x['table'] == 'salebond':
			row += 1
			_date = datetime.datetime.utcfromtimestamp(int(x['input_date'])).strftime('%d.%m.%Y')
			worksheet.write(row, 0, _date)
			worksheet.write(row, 1, 'Продажа облигаций')
			worksheet.write(row, 3, x['ticker'])
			worksheet.write(row, 4, x['count'])
			worksheet.write(row, 2, x['broker'])
			worksheet.write(row, 5, x['price'])
			worksheet.write(row, 7, x['nkd'])
		elif x['table'] == 'taxes':
			row += 1
			_date = datetime.datetime.utcfromtimestamp(int(x['input_date'])).strftime('%d.%m.%Y')
			worksheet.write(row, 0, _date)
			worksheet.write(row, 1, 'Налоги')
			worksheet.write(row, 6, x['amount'])
			worksheet.write(row, 2, x['broker'])
		elif x['table'] == 'comissions':
			row += 1
			_date = datetime.datetime.utcfromtimestamp(int(x['input_date'])).strftime('%d.%m.%Y')
			worksheet.write(row, 0, _date)
			worksheet.write(row, 1, 'Комиссия')
			worksheet.write(row, 6, x['amount'])
			worksheet.write(row, 2, x['broker'])
		elif x['table'] == 'couponincome':
			row += 1
			_date = datetime.datetime.utcfromtimestamp(int(x['input_date'])).strftime('%d.%m.%Y')
			worksheet.write(row, 0, _date)
			worksheet.write(row, 1, 'Купонный доход')
			worksheet.write(row, 6, x['amount'])
			worksheet.write(row, 2, x['broker'])
			worksheet.write(row, 3, x['bond'])
		elif x['table'] == 'dividends':
			row += 1
			_date = datetime.datetime.utcfromtimestamp(int(x['input_date'])).strftime('%d.%m.%Y')
			worksheet.write(row, 0, _date)
			worksheet.write(row, 1, 'Дивиденды')
			worksheet.write(row, 6, x['amount'])
			worksheet.write(row, 2, x['broker'])
			worksheet.write(row, 3, x['dividend'])

	workbook.close()
	return filename


def import_excel_file(uid, filename):
	"""
	Импортировать excel файл с операциями
	"""
	book = openpyxl.load_workbook(filename)
	sheet = book.active

	row = 2
	err_message = ''

	# Перебираем все строки в файле
	while True:
		try:
			operation_type = sheet.cell(row=row, column=2)
			# Проверяем, что ячейка не пустая
			if not operation_type.value:
				break

			# Обработать операции по названию
			if operation_type.value.lower() == 'пополнение счета':
				# _date = int(time.mktime(datetime.datetime.strptime(sheet.cell(row, 1).value, '%d.%m.%Y').timetuple()))
				_date = int(time.mktime((datetime.datetime.strptime(sheet.cell(row, 1).value, '%d.%m.%Y') + datetime.timedelta(hours=3)).timetuple()))
				broker = sheet.cell(row, 3).value
				amount = sheet.cell(row, 7).value
				DataBase.add_new_amount(uid, int(time.time()), _date, amount, broker)
			elif operation_type.value.lower() == 'вывод средств':
				# _date = int(time.mktime(datetime.datetime.strptime(sheet.cell(row, 1).value, '%d.%m.%Y').timetuple()))
				_date = int(time.mktime((datetime.datetime.strptime(sheet.cell(row, 1).value, '%d.%m.%Y') + datetime.timedelta(hours=3)).timetuple()))
				broker = sheet.cell(row, 3).value
				amount = sheet.cell(row, 7).value
				DataBase.add_minus_amount(uid, int(time.time()), _date, amount, broker)
			elif operation_type.value.lower() == 'покупка акций':
				# _date = int(time.mktime(datetime.datetime.strptime(sheet.cell(row, 1).value, '%d.%m.%Y').timetuple()))
				_date = int(time.mktime((datetime.datetime.strptime(sheet.cell(row, 1).value, '%d.%m.%Y') + datetime.timedelta(hours=3)).timetuple()))
				broker = sheet.cell(row, 3).value
				tiker = sheet.cell(row, 4).value
				count = sheet.cell(row, 5).value
				price = sheet.cell(row, 6).value
				api_price = Moex.get_stock_price(tiker.upper())
				if not api_price:
					err_message = 'Строка {!s}: Такого тикера не существует\n'.format(row)
					row += 1
					continue
				DataBase.add_buystock(uid, int(time.time()), _date, tiker, count, broker, price, api_price)
			elif operation_type.value.lower() == 'продажа акций':
				# _date = int(time.mktime(datetime.datetime.strptime(sheet.cell(row, 1).value, '%d.%m.%Y').timetuple()))
				_date = int(time.mktime((datetime.datetime.strptime(sheet.cell(row, 1).value, '%d.%m.%Y') + datetime.timedelta(hours=3)).timetuple()))
				broker = sheet.cell(row, 3).value
				tiker = sheet.cell(row, 4).value
				count = sheet.cell(row, 5).value
				price = sheet.cell(row, 6).value
				api_price = Moex.get_stock_price(tiker.upper())
				if not api_price:
					err_message = 'Строка {!s}: Такого тикера не существует\n'.format(row)
					row += 1
					continue
				DataBase.add_salestock(uid, int(time.time()), _date, tiker, count, broker, price)
			elif operation_type.value.lower() == 'покупка облигаций':
				# _date = int(time.mktime(datetime.datetime.strptime(sheet.cell(row, 1).value, '%d.%m.%Y').timetuple()))
				_date = int(time.mktime((datetime.datetime.strptime(sheet.cell(row, 1).value, '%d.%m.%Y') + datetime.timedelta(hours=3)).timetuple()))
				broker = sheet.cell(row, 3).value
				tiker = sheet.cell(row, 4).value
				count = sheet.cell(row, 5).value
				price = sheet.cell(row, 6).value
				nkd = sheet.cell(row, 8).value
				data = Moex.get_bond_data(tiker.upper())
				if not data:
					err_message = 'Строка {!s}: Такого тикера не существует\n'.format(row)
					row += 1
					continue
				api_price = data['price']
				api_nkd = data['nkd']
				api_FACEVALUE = data['FACEVALUE']
				api_ACCINT = data['ACCINT']
				DataBase.add_buybond(
					uid, int(time.time()), _date, tiker, count, nkd, price, 
					broker, api_price, api_nkd, api_FACEVALUE, api_ACCINT)
			elif operation_type.value.lower() == 'продажа облигаций':
				# _date = int(time.mktime(datetime.datetime.strptime(sheet.cell(row, 1).value, '%d.%m.%Y').timetuple()))
				_date = int(time.mktime((datetime.datetime.strptime(sheet.cell(row, 1).value, '%d.%m.%Y') + datetime.timedelta(hours=3)).timetuple()))
				broker = sheet.cell(row, 3).value
				tiker = sheet.cell(row, 4).value
				count = sheet.cell(row, 5).value
				price = sheet.cell(row, 6).value
				nkd = sheet.cell(row, 8).value
				data = Moex.get_bond_data(tiker.upper())
				if not data:
					err_message = 'Строка {!s}: Такого тикера не существует\n'.format(row)
					row += 1
					continue
				DataBase.add_salebond(uid, int(time.time()), _date, tiker, count, broker, nkd, price)
			elif operation_type.value.lower() in ['налог', 'налоги']:
				# _date = int(time.mktime(datetime.datetime.strptime(sheet.cell(row, 1).value, '%d.%m.%Y').timetuple()))
				_date = int(time.mktime((datetime.datetime.strptime(sheet.cell(row, 1).value, '%d.%m.%Y') + datetime.timedelta(hours=3)).timetuple()))
				broker = sheet.cell(row, 3).value
				amount = sheet.cell(row, 7).value
				DataBase.add_new_tax(uid, int(time.time()), _date, amount, broker)
			elif operation_type.value.lower() in ['комиссия']:
				# _date = int(time.mktime(datetime.datetime.strptime(sheet.cell(row, 1).value, '%d.%m.%Y').timetuple()))
				_date = int(time.mktime((datetime.datetime.strptime(sheet.cell(row, 1).value, '%d.%m.%Y') + datetime.timedelta(hours=3)).timetuple()))
				broker = sheet.cell(row, 3).value
				amount = sheet.cell(row, 7).value
				DataBase.add_new_commission(uid, int(time.time()), _date, amount, broker)
			elif operation_type.value.lower() in ['купонный доход']:
				# _date = int(time.mktime(datetime.datetime.strptime(sheet.cell(row, 1).value, '%d.%m.%Y').timetuple()))
				_date = int(time.mktime((datetime.datetime.strptime(sheet.cell(row, 1).value, '%d.%m.%Y') + datetime.timedelta(hours=3)).timetuple()))
				broker = sheet.cell(row, 3).value
				tiker = sheet.cell(row, 4).value
				amount = sheet.cell(row, 7).value
				DataBase.add_new_couponincome(uid, int(time.time()), _date, tiker, amount, broker)
			elif operation_type.value.lower() in ['дивиденды']:
				# _date = int(time.mktime(datetime.datetime.strptime(sheet.cell(row, 1).value, '%d.%m.%Y').timetuple()))
				_date = int(time.mktime((datetime.datetime.strptime(sheet.cell(row, 1).value, '%d.%m.%Y') + datetime.timedelta(hours=3)).timetuple()))
				broker = sheet.cell(row, 3).value
				tiker = sheet.cell(row, 4).value
				amount = sheet.cell(row, 7).value
				DataBase.add_new_dividend(uid, int(time.time()), _date, tiker, amount, broker)
		except Exception as e:
			print(e)
			err_message += 'Ошибка в строке {!s}\n'.format(row)
			continue

		row += 1

	return err_message


# print(import_excel_file(217166737, 'import_217166737.xlsx'))
# print(get_account_state(217166737))
# print(get_portfolio(217166737))
# print(get_timestamp('21.03.3000'))
# print(Moex.get_stock_price('HYDR'))
# print(Moex.get_bond_data('SU26204RMFS6'))  # SU26210RMFS3
# print(get_portfolio_amount(217166737))
# update_moex()
